from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def excel_report(df, stats_df, correlation_matrix):

    wb = Workbook()

    wb.remove(wb.active)

    ws1 = wb.create_sheet(title="Raw_Data")

    df = df.copy()
    df["Date"] = df["Date"].dt.date

    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    
    format_sheet(ws1)
    auto_adjust_column_width(ws1)

    ws2 = wb.create_sheet(title="Summary_Stats")

    summ_stats_df = stats_df.reset_index()

    for r in dataframe_to_rows(summ_stats_df, index=True, header=True):
        ws2.append(r)

    format_sheet(ws2)
    auto_adjust_column_width(ws2)

    ws3 = wb.create_sheet(title="Correlation")

    for r in dataframe_to_rows(correlation_matrix, index=True, header=True):
        ws3.append(r)

    format_sheet(ws3)
    auto_adjust_column_width(ws3)
        
    ws4 = wb.create_sheet(title="RSI_Last_30Days")

    rsi_df = df.sort_values("Date").groupby("Ticker").tail(30)

    for ticker, group in rsi_df.groupby("Ticker"):

        ws4.append([ticker])  # ticker header

        ws4.append(["Date", "RSI"])  # column headers

        for _, row in group.iterrows():
            ws4.append([row["Date"], row["RSI"]])

        ws4.append([])  # spacing

    format_sheet(ws4)
    auto_adjust_column_width(ws4)

    headers = [cell.value for cell in ws1[1]]
    return_idx = headers.index("Return")

    apply_return_formatting(ws1, return_idx)

    wb.save("data/stock_analysis_report.xlsx")

def format_sheet(ws):

    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    ws.freeze_panes = "A2"

def auto_adjust_column_width(ws):

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max_length + 2

def apply_return_formatting(ws, return_col_index):

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2):

        cell = row[return_col_index]

        if cell.value is None:
            continue

        if cell.value > 0:
            cell.fill = green_fill
        elif cell.value < 0:
            cell.fill = red_fill
